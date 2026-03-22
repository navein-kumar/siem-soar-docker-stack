"""Inventory Excel Export - Multi-sheet workbook with raw data per agent"""
import os
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import config, opensearch_client

IST = timezone(timedelta(hours=5, minutes=30))

# Styles
HEADER_FONT = Font(name='Segoe UI', bold=True, size=10, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
DATA_FONT = Font(name='Segoe UI', size=9)
ALT_FILL = PatternFill(start_color='F8F9FB', end_color='F8F9FB', fill_type='solid')
TITLE_FONT = Font(name='Segoe UI', bold=True, size=14, color='1B2A4A')
SUBTITLE_FONT = Font(name='Segoe UI', size=10, color='69707D')
BORDER = Border(
    bottom=Side(style='thin', color='E5E7EB')
)

def _add_header(ws, row, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER

def _add_row(ws, row, values, alt=False):
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.font = DATA_FONT
        cell.border = BORDER
        if alt:
            cell.fill = ALT_FILL

def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

def _query_raw(index_pattern, source_fields, agent_filter=None, size=10000):
    """Get raw documents from an inventory index. Uses scroll for >10K docs."""
    client = opensearch_client.get_client()
    body = {"size": min(size, 10000), "_source": source_fields}
    if agent_filter:
        body["query"] = {"term": {"agent.name": agent_filter}}

    # First batch
    result = client.search(index=index_pattern, body=body, scroll="2m")
    scroll_id = result.get("_scroll_id")
    hits = [h["_source"] for h in result["hits"]["hits"]]
    total = result["hits"]["total"]["value"]

    # Scroll for remaining if needed
    while len(hits) < total and len(hits) < size and scroll_id:
        result = client.scroll(scroll_id=scroll_id, scroll="2m")
        batch = [h["_source"] for h in result["hits"]["hits"]]
        if not batch:
            break
        hits.extend(batch)
        scroll_id = result.get("_scroll_id")

    # Clear scroll
    if scroll_id:
        try:
            client.clear_scroll(scroll_id=scroll_id)
        except:
            pass

    return hits

def _get_nested(obj, path, default=""):
    """Get nested value from dict using dot notation"""
    keys = path.split(".")
    for k in keys:
        if isinstance(obj, dict):
            obj = obj.get(k, default)
        else:
            return default
    return obj if obj is not None else default

def generate_inventory_excel(agent_filter=None):
    """Generate multi-sheet Excel inventory workbook"""
    wb = Workbook()
    idx_prefix = config.OPENSEARCH_INDEX.split('-')[0]
    now = datetime.now(IST)
    agent_label = agent_filter or "All Agents"

    # ============================================================
    # Sheet 1: Summary
    # ============================================================
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value="IT Asset Inventory Report").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Generated: {now.strftime('%d/%m/%Y %I:%M %p IST')} | Agent: {agent_label}").font = SUBTITLE_FONT
    ws.cell(row=3, column=1, value="")

    # Get summary counts
    client = opensearch_client.get_client()
    summary_data = []
    for label, idx in [
        ("Endpoints (System)", "system"), ("Hardware", "hardware"),
        ("Packages", "packages"), ("Processes", "processes"),
        ("Ports/Listeners", "ports"), ("Services", "services"),
        ("Users", "users"), ("Browser Extensions", "browser-extensions"),
        ("Interfaces", "interfaces"), ("Networks", "networks"),
        ("Hotfixes", "hotfixes"), ("Vulnerabilities (states)", "vulnerabilities")
    ]:
        try:
            idx_name = f"{idx_prefix}-states-inventory-{idx}-*" if idx != "vulnerabilities" else f"{idx_prefix}-states-{idx}-*"
            body = {"size": 0}
            if agent_filter:
                body["query"] = {"term": {"agent.name": agent_filter}}
            r = client.search(index=idx_name, body=body)
            summary_data.append((label, r["hits"]["total"]["value"]))
        except:
            summary_data.append((label, 0))

    _add_header(ws, 5, ["Category", "Count"])
    for i, (label, count) in enumerate(summary_data):
        _add_row(ws, 6+i, [label, count], i % 2 == 1)
    _auto_width(ws)

    # ============================================================
    # Sheet 2: Hardware
    # ============================================================
    ws2 = wb.create_sheet("Hardware")
    try:
        docs = _query_raw(f"{idx_prefix}-states-inventory-hardware-*",
            ["agent.name", "host.cpu.name", "host.cpu.cores", "host.memory.total", "host.memory.used", "host.memory.usage", "host.serial_number"],
            agent_filter, 100)
        _add_header(ws2, 1, ["Agent", "CPU", "Cores", "RAM Total (GB)", "RAM Used (GB)", "Usage %", "Serial Number"])
        for i, d in enumerate(docs):
            mem_total = _get_nested(d, "host.memory.total", 0)
            mem_used = _get_nested(d, "host.memory.used", 0)
            usage = _get_nested(d, "host.memory.usage", 0)
            _add_row(ws2, 2+i, [
                _get_nested(d, "agent.name"),
                _get_nested(d, "host.cpu.name"),
                _get_nested(d, "host.cpu.cores"),
                round(mem_total / (1024**3), 1) if mem_total else 0,
                round(mem_used / (1024**3), 1) if mem_used else 0,
                round(usage * 100, 1) if isinstance(usage, float) and usage <= 1 else usage,
                _get_nested(d, "host.serial_number")
            ], i % 2 == 1)
    except:
        ws2.cell(row=1, column=1, value="No hardware data available")
    _auto_width(ws2)

    # ============================================================
    # Sheet 3: Software/Packages
    # ============================================================
    ws3 = wb.create_sheet("Packages")
    try:
        docs = _query_raw(f"{idx_prefix}-states-inventory-packages-*",
            ["agent.name", "package.name", "package.version", "package.vendor", "package.architecture", "package.type", "package.size"],
            agent_filter, 100000)
        _add_header(ws3, 1, ["Agent", "Package Name", "Version", "Vendor", "Architecture", "Type", "Size"])
        for i, d in enumerate(docs):
            _add_row(ws3, 2+i, [
                _get_nested(d, "agent.name"),
                _get_nested(d, "package.name"),
                _get_nested(d, "package.version"),
                _get_nested(d, "package.vendor"),
                _get_nested(d, "package.architecture"),
                _get_nested(d, "package.type"),
                _get_nested(d, "package.size", 0)
            ], i % 2 == 1)
    except:
        ws3.cell(row=1, column=1, value="No package data available")
    _auto_width(ws3)

    # ============================================================
    # Sheet 4: Processes
    # ============================================================
    ws4 = wb.create_sheet("Processes")
    try:
        docs = _query_raw(f"{idx_prefix}-states-inventory-processes-*",
            ["agent.name", "process.name", "process.pid", "process.command_line", "process.start"],
            agent_filter, 10000)
        _add_header(ws4, 1, ["Agent", "Process Name", "PID", "Command Line", "Start Time"])
        for i, d in enumerate(docs):
            _add_row(ws4, 2+i, [
                _get_nested(d, "agent.name"),
                _get_nested(d, "process.name"),
                _get_nested(d, "process.pid"),
                str(_get_nested(d, "process.command_line", ""))[:100],
                _get_nested(d, "process.start")
            ], i % 2 == 1)
    except:
        ws4.cell(row=1, column=1, value="No process data available")
    _auto_width(ws4)

    # ============================================================
    # Sheet 5: Ports & Listeners
    # ============================================================
    ws5 = wb.create_sheet("Ports & Listeners")
    try:
        docs = _query_raw(f"{idx_prefix}-states-inventory-ports-*",
            ["agent.name", "source.ip", "source.port", "destination.ip", "destination.port", "network.transport", "process.name", "process.pid", "interface.state"],
            agent_filter, 10000)
        _add_header(ws5, 1, ["Agent", "Source IP", "Source Port", "Dest IP", "Dest Port", "Protocol", "Process", "PID", "State"])
        for i, d in enumerate(docs):
            _add_row(ws5, 2+i, [
                _get_nested(d, "agent.name"),
                _get_nested(d, "source.ip"),
                _get_nested(d, "source.port"),
                _get_nested(d, "destination.ip"),
                _get_nested(d, "destination.port"),
                _get_nested(d, "network.transport"),
                _get_nested(d, "process.name"),
                _get_nested(d, "process.pid"),
                _get_nested(d, "interface.state")
            ], i % 2 == 1)
    except:
        ws5.cell(row=1, column=1, value="No port data available")
    _auto_width(ws5)

    # ============================================================
    # Sheet 6: Services
    # ============================================================
    ws6 = wb.create_sheet("Services")
    try:
        docs = _query_raw(f"{idx_prefix}-states-inventory-services-*",
            ["agent.name", "service.name", "service.state", "service.type", "service.start_type", "process.executable"],
            agent_filter, 10000)
        _add_header(ws6, 1, ["Agent", "Service Name", "State", "Type", "Start Type", "Executable"])
        for i, d in enumerate(docs):
            _add_row(ws6, 2+i, [
                _get_nested(d, "agent.name"),
                _get_nested(d, "service.name"),
                _get_nested(d, "service.state"),
                _get_nested(d, "service.type"),
                _get_nested(d, "service.start_type"),
                str(_get_nested(d, "process.executable", ""))[:80]
            ], i % 2 == 1)
    except:
        ws6.cell(row=1, column=1, value="No service data available")
    _auto_width(ws6)

    # ============================================================
    # Sheet 7: Users
    # ============================================================
    ws7 = wb.create_sheet("Users")
    try:
        docs = _query_raw(f"{idx_prefix}-states-inventory-users-*",
            ["agent.name", "user.name", "user.type", "user.groups", "user.shell", "user.full_name", "login.status"],
            agent_filter, 5000)
        _add_header(ws7, 1, ["Agent", "Username", "Full Name", "Type", "Groups", "Shell", "Logged In"])
        for i, d in enumerate(docs):
            groups = _get_nested(d, "user.groups", [])
            if isinstance(groups, list):
                groups = ", ".join(groups)
            _add_row(ws7, 2+i, [
                _get_nested(d, "agent.name"),
                _get_nested(d, "user.name"),
                _get_nested(d, "user.full_name"),
                _get_nested(d, "user.type"),
                str(groups),
                _get_nested(d, "user.shell"),
                "Yes" if _get_nested(d, "login.status") else "No"
            ], i % 2 == 1)
    except:
        ws7.cell(row=1, column=1, value="No user data available")
    _auto_width(ws7)

    # ============================================================
    # Sheet 8: Browser Extensions
    # ============================================================
    ws8 = wb.create_sheet("Browser Extensions")
    try:
        docs = _query_raw(f"{idx_prefix}-states-inventory-browser-extensions-*",
            ["agent.name", "browser.name", "package.name", "package.version", "package.vendor", "package.enabled", "package.type"],
            agent_filter, 5000)
        _add_header(ws8, 1, ["Agent", "Browser", "Extension Name", "Version", "Vendor", "Enabled", "Type"])
        for i, d in enumerate(docs):
            _add_row(ws8, 2+i, [
                _get_nested(d, "agent.name"),
                _get_nested(d, "browser.name"),
                _get_nested(d, "package.name"),
                _get_nested(d, "package.version"),
                _get_nested(d, "package.vendor"),
                "Yes" if _get_nested(d, "package.enabled") else "No",
                _get_nested(d, "package.type")
            ], i % 2 == 1)
    except:
        ws8.cell(row=1, column=1, value="No browser extension data available")
    _auto_width(ws8)

    # ============================================================
    # Sheet 9: Network
    # ============================================================
    ws9 = wb.create_sheet("Network")
    try:
        docs = _query_raw(f"{idx_prefix}-states-inventory-networks-*",
            ["agent.name", "interface.name", "network.ip", "network.netmask", "network.broadcast", "network.type"],
            agent_filter, 1000)
        _add_header(ws9, 1, ["Agent", "Interface", "IP Address", "Netmask", "Broadcast", "Type"])
        for i, d in enumerate(docs):
            _add_row(ws9, 2+i, [
                _get_nested(d, "agent.name"),
                _get_nested(d, "interface.name"),
                _get_nested(d, "network.ip"),
                _get_nested(d, "network.netmask"),
                _get_nested(d, "network.broadcast"),
                _get_nested(d, "network.type")
            ], i % 2 == 1)
    except:
        ws9.cell(row=1, column=1, value="No network data available")
    _auto_width(ws9)

    # ============================================================
    # Sheet 10: Hotfixes
    # ============================================================
    ws10 = wb.create_sheet("Hotfixes")
    try:
        docs = _query_raw(f"{idx_prefix}-states-inventory-hotfixes-*",
            ["agent.name", "package.hotfix.name"],
            agent_filter, 5000)
        _add_header(ws10, 1, ["Agent", "Hotfix (KB)"])
        for i, d in enumerate(docs):
            _add_row(ws10, 2+i, [
                _get_nested(d, "agent.name"),
                _get_nested(d, "package.hotfix.name")
            ], i % 2 == 1)
    except:
        ws10.cell(row=1, column=1, value="No hotfix data available")
    _auto_width(ws10)

    # ============================================================
    # Sheet 11: Vulnerabilities
    # ============================================================
    ws11 = wb.create_sheet("Vulnerabilities")
    try:
        docs = _query_raw(f"{idx_prefix}-states-vulnerabilities-*",
            ["agent.name", "vulnerability.id", "vulnerability.severity", "vulnerability.description",
             "vulnerability.score.base", "package.name", "package.version", "vulnerability.detected_at",
             "vulnerability.reference", "vulnerability.category"],
            agent_filter, 100000)
        _add_header(ws11, 1, ["Agent", "CVE ID", "Severity", "CVSS Score", "Category", "Package", "Version", "Description", "Detected At", "Reference"])
        for i, d in enumerate(docs):
            desc = str(_get_nested(d, "vulnerability.description", ""))
            if len(desc) > 100:
                desc = desc[:100] + "..."
            _add_row(ws11, 2+i, [
                _get_nested(d, "agent.name"),
                _get_nested(d, "vulnerability.id"),
                _get_nested(d, "vulnerability.severity"),
                _get_nested(d, "vulnerability.score.base"),
                _get_nested(d, "vulnerability.category"),
                _get_nested(d, "package.name"),
                _get_nested(d, "package.version"),
                desc,
                _get_nested(d, "vulnerability.detected_at"),
                _get_nested(d, "vulnerability.reference")
            ], i % 2 == 1)
    except:
        ws11.cell(row=1, column=1, value="No vulnerability data available")
    _auto_width(ws11)

    # Save
    fname = f"inventory_{agent_label.replace(' ','_')}_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    fpath = os.path.join(config.REPORTS_DIR, fname)
    os.makedirs(config.REPORTS_DIR, exist_ok=True)
    wb.save(fpath)

    return {"filename": fname, "filepath": fpath, "size": os.path.getsize(fpath)}
