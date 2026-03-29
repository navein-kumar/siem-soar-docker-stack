"""Incident Management Excel Export — TheHive data, multi-sheet workbook.
Uses identical styles/helpers as security_excel.py for consistent formatting."""

from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io, thehive_client

IST = timezone(timedelta(hours=5, minutes=30))

# ── Identical styles to security_excel.py ───────────────────
HEADER_FONT  = Font(name='Segoe UI', bold=True, size=10, color='FFFFFF')
HEADER_FILL  = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
DATA_FONT    = Font(name='Segoe UI', size=9)
ALT_FILL     = PatternFill(start_color='F8F9FB', end_color='F8F9FB', fill_type='solid')
TITLE_FONT   = Font(name='Segoe UI', bold=True, size=14, color='1B2A4A')
SUBTITLE_FONT= Font(name='Segoe UI', size=10, color='69707D')
BOLD_FONT    = Font(name='Segoe UI', bold=True, size=9, color='1B2A4A')
BORDER       = Border(bottom=Side(style='thin', color='E5E7EB'))
WARN_FILL    = PatternFill(start_color='FEF0EB', end_color='FEF0EB', fill_type='solid')
CRIT_FILL    = PatternFill(start_color='FDECEA', end_color='FDECEA', fill_type='solid')

SEV_COLORS = {
    'Critical': PatternFill(start_color='BD271E', end_color='BD271E', fill_type='solid'),
    'High':     PatternFill(start_color='E7664C', end_color='E7664C', fill_type='solid'),
    'Medium':   PatternFill(start_color='D6BF57', end_color='D6BF57', fill_type='solid'),
    'Low':      PatternFill(start_color='6DCCB1', end_color='6DCCB1', fill_type='solid'),
}


def _add_header(ws, row, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER


def _add_row(ws, row, values, alt=False, sev=None):
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.font = DATA_FONT
        cell.border = BORDER
        if sev and col == 1:
            fill = SEV_COLORS.get(sev)
            if fill:
                cell.fill = fill
                cell.font = Font(name='Segoe UI', bold=True, size=9, color='FFFFFF')
        elif alt:
            cell.fill = ALT_FILL


def _auto_width(ws, max_width=55):
    for col in ws.columns:
        mx = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 3, max_width)


def _title(ws, title, subtitle):
    ws.merge_cells('A1:H1')
    ws['A1'].value = title
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 28
    ws.merge_cells('A2:H2')
    ws['A2'].value = subtitle
    ws['A2'].font = SUBTITLE_FONT
    ws['A2'].alignment = Alignment(horizontal='left')


# ── Sheet 1: Summary ────────────────────────────────────────
def _sheet_summary(wb, d):
    ws = wb.active
    ws.title = 'Summary'
    ws.sheet_view.showGridLines = False

    gen_time = datetime.now(IST).strftime('%d/%m/%Y %H:%M IST')
    _title(ws, 'Incident Management Report — Summary', f'Generated: {gen_time}')

    row = 4
    _add_header(ws, row, ['Metric', 'Value'])
    row += 1
    kpis = [
        ('Total Alerts', d['total_alerts']),
        ('Critical Alerts', d['alert_sev'].get('Critical', 0)),
        ('High Alerts', d['alert_sev'].get('High', 0)),
        ('Medium Alerts', d['alert_sev'].get('Medium', 0)),
        ('Low Alerts', d['alert_sev'].get('Low', 0)),
        ('Total Cases', d['total_cases']),
        ('Open Cases', d['open_cases_count']),
        ('Closed Cases', d['closed_cases_count']),
        ('Mean Time to Resolve', d['mttr_display']),
        ('True Positive Rate', f"{d['tp_rate']}%"),
        ('SLA Breaches >72h', d['sla_72h']),
        ('Cases >24h Open', d['sla_24h']),
        ('Total Observables', sum(d['obs_summary'].values())),
    ]
    for i, (k, v) in enumerate(kpis):
        alt = i % 2 == 0
        _add_row(ws, row + i, [k, v], alt=alt)
        if 'Critical' in k:
            ws.cell(row+i, 1).fill = SEV_COLORS['Critical']
            ws.cell(row+i, 1).font = Font(name='Segoe UI', bold=True, size=9, color='FFFFFF')
        elif 'High' in k:
            ws.cell(row+i, 1).fill = SEV_COLORS['High']
            ws.cell(row+i, 1).font = Font(name='Segoe UI', bold=True, size=9, color='FFFFFF')
        elif 'Medium' in k:
            ws.cell(row+i, 1).fill = SEV_COLORS['Medium']
            ws.cell(row+i, 1).font = Font(name='Segoe UI', bold=True, size=9, color='FFFFFF')

    # Case severity breakdown (right side)
    ws.cell(4, 4).value = 'ALERT SEVERITY'
    ws.cell(4, 4).font = BOLD_FONT
    _add_header(ws, 5, ['', 'Severity', 'Alerts', 'Cases', ''])
    sevs = ['Critical', 'High', 'Medium', 'Low']
    for i, s in enumerate(sevs):
        r = 6 + i
        ws.cell(r, 4).value = ''
        ws.cell(r, 5).value = s
        ws.cell(r, 5).font = Font(name='Segoe UI', bold=True, size=9, color='FFFFFF')
        ws.cell(r, 5).fill = SEV_COLORS[s]
        ws.cell(r, 6).value = d['alert_sev'].get(s, 0)
        ws.cell(r, 6).font = BOLD_FONT
        ws.cell(r, 7).value = d['case_sev'].get(s, 0)
        ws.cell(r, 7).font = BOLD_FONT

    # MITRE top
    ws.cell(12, 4).value = 'TOP MITRE TECHNIQUES'
    ws.cell(12, 4).font = BOLD_FONT
    _add_header(ws, 13, ['', 'Technique', 'Count', '', ''])
    for i, (t, c) in enumerate(d['top_mitre'][:8]):
        ws.cell(14+i, 5).value = t
        ws.cell(14+i, 6).value = c
        ws.cell(14+i, 5).font = DATA_FONT
        ws.cell(14+i, 6).font = BOLD_FONT
        if i % 2 == 0:
            ws.cell(14+i, 5).fill = ALT_FILL
            ws.cell(14+i, 6).fill = ALT_FILL

    _auto_width(ws)


# ── Sheet 2: Alerts ─────────────────────────────────────────
def _sheet_alerts(wb, d):
    ws = wb.create_sheet('Alerts')
    ws.sheet_view.showGridLines = False
    _title(ws, 'Alert Analysis', f'Top triggered alerts from TheHive — {d["total_alerts"]} total')

    row = 4
    _add_header(ws, row, ['#', 'Alert Title', 'Count'])
    for i, (title, cnt) in enumerate(d['top_alert_titles'], 1):
        _add_row(ws, row+i, [i, title, cnt], alt=i % 2 == 0)
    row += len(d['top_alert_titles']) + 2

    ws.cell(row, 1).value = 'ALERT STATUS BREAKDOWN'
    ws.cell(row, 1).font = BOLD_FONT
    _add_header(ws, row+1, ['Status', 'Count'])
    for i, (status, cnt) in enumerate(d['alert_status'].items()):
        _add_row(ws, row+2+i, [status, cnt], alt=i % 2 == 0)
    row += len(d['alert_status']) + 4

    ws.cell(row, 1).value = 'TOP TAGS'
    ws.cell(row, 1).font = BOLD_FONT
    _add_header(ws, row+1, ['Tag', 'Count'])
    for i, (tag, cnt) in enumerate(d['top_tags']):
        _add_row(ws, row+2+i, [tag, cnt], alt=i % 2 == 0)

    _auto_width(ws)


# ── Sheet 3: Open Cases ──────────────────────────────────────
def _sheet_cases(wb, d):
    ws = wb.create_sheet('Open Cases')
    ws.sheet_view.showGridLines = False
    _title(ws, 'Open Cases', f'{d["open_cases_count"]} active cases — {d["closed_cases_count"]} closed')

    _add_header(ws, 4, ['Severity', 'Case #', 'Title', 'Status', 'Assignee', 'Age (hrs)', 'Tags', 'SLA Status'])
    for i, c in enumerate(d['open_cases']):
        r = 5 + i
        alt = i % 2 == 0
        sla = 'BREACH >72h' if c['age_h'] > 72 else ('WARNING >24h' if c['age_h'] > 24 else 'OK')
        tags_str = ', '.join(c.get('tags', []))
        _add_row(ws, r, [c['severity'], f"#{c.get('number','')}", c['title'],
                         c['status'], c['assignee'], round(c['age_h'], 1), tags_str, sla],
                 alt=alt, sev=c['severity'])
        sla_cell = ws.cell(row=r, column=8)
        if c['age_h'] > 72:
            sla_cell.fill = CRIT_FILL
            sla_cell.font = Font(name='Segoe UI', bold=True, size=9, color='BD271E')
        elif c['age_h'] > 24:
            sla_cell.fill = WARN_FILL
            sla_cell.font = Font(name='Segoe UI', bold=True, size=9, color='E7664C')
        else:
            sla_cell.font = Font(name='Segoe UI', bold=True, size=9, color='27AE60')

    # Resolution breakdown below
    row = len(d['open_cases']) + 7
    ws.cell(row, 1).value = 'CASE RESOLUTION BREAKDOWN'
    ws.cell(row, 1).font = BOLD_FONT
    _add_header(ws, row+1, ['Resolution Type', 'Count', 'Percentage'])
    total_closed = d['closed_cases_count'] or 1
    for i, (status, cnt) in enumerate(d['resolution_counts'].items()):
        pct = f"{round(cnt/total_closed*100)}%"
        _add_row(ws, row+2+i, [status, cnt, pct], alt=i % 2 == 0)

    _auto_width(ws)


# ── Sheet 4: Observables / IOCs ──────────────────────────────
def _sheet_observables(wb, d):
    ws = wb.create_sheet('Observables IOCs')
    ws.sheet_view.showGridLines = False
    _title(ws, 'Observables / Indicators of Compromise', 'Extracted from TheHive case observables')

    ws.cell(4, 1).value = 'OBSERVABLE TYPES SUMMARY'
    ws.cell(4, 1).font = BOLD_FONT
    _add_header(ws, 5, ['Type', 'Count'])
    for i, (t, c) in enumerate(sorted(d['obs_summary'].items(), key=lambda x: -x[1])):
        _add_row(ws, 6+i, [t, c], alt=i % 2 == 0)

    ws.cell(4, 3).value = 'TOP SOURCE IPs'
    ws.cell(4, 3).font = BOLD_FONT
    _add_header(ws, 5, ['', 'IP Address', 'Count', ''])
    for i, (ip, cnt) in enumerate(d['top_ips']):
        ws.cell(6+i, 3).value = ''
        ws.cell(6+i, 4).value = ip
        ws.cell(6+i, 4).font = Font(name='Courier New', size=9)
        ws.cell(6+i, 5).value = cnt
        ws.cell(6+i, 5).font = BOLD_FONT
        if i % 2 == 0:
            ws.cell(6+i, 4).fill = ALT_FILL
            ws.cell(6+i, 5).fill = ALT_FILL

    ws.cell(4, 7).value = 'TOP HOSTNAMES'
    ws.cell(4, 7).font = BOLD_FONT
    _add_header(ws, 5, ['', 'Hostname', 'Count', ''])
    for i, (h, cnt) in enumerate(d['top_hostnames']):
        ws.cell(6+i, 7).value = ''
        ws.cell(6+i, 8).value = h
        ws.cell(6+i, 8).font = Font(name='Courier New', size=9)
        ws.cell(6+i, 9).value = cnt
        if i % 2 == 0:
            ws.cell(6+i, 8).fill = ALT_FILL
            ws.cell(6+i, 9).fill = ALT_FILL

    hash_start = max(len(d['obs_summary']), len(d['top_ips'])) + 8
    ws.cell(hash_start, 1).value = 'FILE HASHES'
    ws.cell(hash_start, 1).font = BOLD_FONT
    _add_header(ws, hash_start+1, ['Hash', 'Count'])
    for i, (h, cnt) in enumerate(d['top_hashes']):
        ws.cell(hash_start+2+i, 1).value = h
        ws.cell(hash_start+2+i, 1).font = Font(name='Courier New', size=8)
        ws.cell(hash_start+2+i, 2).value = cnt
        if i % 2 == 0:
            ws.cell(hash_start+2+i, 1).fill = ALT_FILL
            ws.cell(hash_start+2+i, 2).fill = ALT_FILL

    _auto_width(ws)


# ── Sheet 5: MITRE ATT&CK ────────────────────────────────────
def _sheet_mitre(wb, d):
    ws = wb.create_sheet('MITRE ATT&CK')
    ws.sheet_view.showGridLines = False
    _title(ws, 'MITRE ATT&CK Techniques Detected', 'Based on alert tags from TheHive/Wazuh integration')

    _add_header(ws, 4, ['#', 'Technique / Tactic', 'Alert Count', 'Risk Level'])
    for i, (technique, cnt) in enumerate(d['top_mitre'], 1):
        risk = 'Critical' if cnt > 20 else ('High' if cnt > 10 else ('Medium' if cnt > 5 else 'Low'))
        _add_row(ws, 4+i, [i, technique, cnt, risk], alt=i % 2 == 0)
        risk_cell = ws.cell(row=4+i, column=4)
        risk_cell.fill = SEV_COLORS.get(risk, ALT_FILL)
        risk_cell.font = Font(name='Segoe UI', bold=True, size=9, color='FFFFFF')

    if not d['top_mitre']:
        ws.cell(5, 1).value = 'No MITRE data yet. Ensure Wazuh alerts include MITRE tags.'
        ws.cell(5, 1).font = Font(name='Segoe UI', italic=True, size=9, color='95A5A6')

    _auto_width(ws)


# ── Sheet 6: Analyst Activity ────────────────────────────────
def _sheet_analysts(wb, d):
    ws = wb.create_sheet('Analyst Activity')
    ws.sheet_view.showGridLines = False
    _title(ws, 'Analyst Activity', 'Case assignment and workload distribution')

    _add_header(ws, 4, ['Analyst', 'Cases Assigned', '% of Total', 'Workload'])
    total = d['total_cases'] or 1
    for i, (analyst, cnt) in enumerate(sorted(d['case_assignees'].items(), key=lambda x: -x[1])):
        pct = round(cnt / total * 100)
        load = 'High' if pct > 40 else ('Medium' if pct > 20 else 'Normal')
        _add_row(ws, 5+i, [analyst, cnt, f"{pct}%", load], alt=i % 2 == 0)

    _auto_width(ws)


# ── Sheet 7: Timeline ────────────────────────────────────────
def _sheet_timeline(wb, d):
    ws = wb.create_sheet('Timeline')
    ws.sheet_view.showGridLines = False
    _title(ws, '7-Day Activity Timeline', 'Alert and case creation over the last 7 days')

    _add_header(ws, 4, ['Date', 'Alerts', 'Cases'])
    for i, (at, ct) in enumerate(zip(d['alert_timeline'], d['case_timeline'])):
        _add_row(ws, 5+i, [at['label'], at['value'], ct['value']], alt=i % 2 == 0)

    _auto_width(ws)


# ── Sheet 8: Recommendations ────────────────────────────────
def _sheet_recs(wb, d):
    ws = wb.create_sheet('Recommendations')
    ws.sheet_view.showGridLines = False
    _title(ws, 'Security Recommendations', 'Auto-generated based on current incident data')

    recs = []
    if d['alert_sev'].get('Critical', 0) > 0:
        recs.append(('Critical', 'Critical Alert Response', f"{d['alert_sev']['Critical']} critical alerts detected. Escalate immediately to Tier-2/3 analysts."))
    if d['sla_72h'] > 0:
        recs.append(('High', 'SLA Breach — Overdue Cases', f"{d['sla_72h']} case(s) exceed 72h SLA. Assign or escalate immediately."))
    if d['sla_24h'] > 0:
        recs.append(('Medium', 'Pending Triage', f"{d['sla_24h']} case(s) pending over 24h. Review and assign to available analysts."))
    if d['tp_rate'] < 50 and d['closed_cases_count'] > 3:
        recs.append(('Medium', 'False Positive Reduction', f"True positive rate is {d['tp_rate']}%. Review alert tuning rules."))
    if d['top_mitre']:
        recs.append(('High', f"Top MITRE: {d['top_mitre'][0][0]}", 'Most frequently detected technique. Review detection coverage.'))
    if d['top_ips']:
        recs.append(('Medium', 'Repeat Offender IP', f"IP {d['top_ips'][0][0]} most frequent in observables. Consider blocking."))
    if not recs:
        recs.append(('Low', 'All Clear', 'No critical recommendations. Continue monitoring and periodic case reviews.'))

    _add_header(ws, 4, ['Priority', 'Recommendation', 'Action Required'])
    for i, (prio, title, action) in enumerate(recs):
        r = 5 + i
        ws.cell(r, 1).value = prio
        ws.cell(r, 1).fill = SEV_COLORS.get(prio, ALT_FILL)
        ws.cell(r, 1).font = Font(name='Segoe UI', bold=True, size=9, color='FFFFFF')
        ws.cell(r, 1).border = BORDER
        ws.cell(r, 2).value = title
        ws.cell(r, 2).font = BOLD_FONT
        ws.cell(r, 2).border = BORDER
        if i % 2 == 0:
            ws.cell(r, 2).fill = ALT_FILL
        ws.cell(r, 3).value = action
        ws.cell(r, 3).font = DATA_FONT
        ws.cell(r, 3).alignment = Alignment(wrap_text=True)
        ws.cell(r, 3).border = BORDER
        if i % 2 == 0:
            ws.cell(r, 3).fill = ALT_FILL

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 70
    for i in range(len(recs)):
        ws.row_dimensions[5+i].height = 30


# ── Main entry ───────────────────────────────────────────────
async def generate_incident_excel(period="7d"):
    d = await thehive_client.fetch_all(period)
    wb = Workbook()

    _sheet_summary(wb, d)
    _sheet_alerts(wb, d)
    _sheet_cases(wb, d)
    _sheet_observables(wb, d)
    _sheet_mitre(wb, d)
    _sheet_analysts(wb, d)
    _sheet_timeline(wb, d)
    _sheet_recs(wb, d)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
