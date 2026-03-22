"""Security Events Excel Export - Multi-sheet workbook with alert data.
Supports: security events, authentication, vulnerability, MITRE, compliance.
Uses scroll API for unlimited data export."""

import os
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import config, opensearch_client

IST = timezone(timedelta(hours=5, minutes=30))
MAX_ROWS = int(os.getenv("EXCEL_MAX_ROWS", "200000"))

# Styles
HEADER_FONT = Font(name='Segoe UI', bold=True, size=10, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
DATA_FONT = Font(name='Segoe UI', size=9)
ALT_FILL = PatternFill(start_color='F8F9FB', end_color='F8F9FB', fill_type='solid')
TITLE_FONT = Font(name='Segoe UI', bold=True, size=14, color='1B2A4A')
SUBTITLE_FONT = Font(name='Segoe UI', size=10, color='69707D')
SEV_COLORS = {
    'Critical': PatternFill(start_color='BD271E', end_color='BD271E', fill_type='solid'),
    'High': PatternFill(start_color='E7664C', end_color='E7664C', fill_type='solid'),
    'Medium': PatternFill(start_color='D6BF57', end_color='D6BF57', fill_type='solid'),
    'Low': PatternFill(start_color='6DCCB1', end_color='6DCCB1', fill_type='solid'),
}
BORDER = Border(bottom=Side(style='thin', color='E5E7EB'))


def _add_header(ws, row, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER


def _add_row(ws, row, values, alt=False):
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.font = DATA_FONT
        cell.border = BORDER
        if alt:
            cell.fill = ALT_FILL


def _auto_width(ws, max_width=50):
    for col in ws.columns:
        mx = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                mx = max(mx, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(mx + 3, max_width)


def _severity(level):
    if level >= 15: return 'Critical'
    if level >= 12: return 'High'
    if level >= 7: return 'Medium'
    return 'Low'


def _scroll_alerts(query, fields, max_rows=MAX_ROWS, progress_cb=None, progress_label=""):
    """Scroll through wazuh-alerts-* index and yield rows"""
    client = opensearch_client.get_client()
    body = {"size": 5000, "query": query, "_source": fields, "sort": [{"timestamp": "desc"}]}
    r = client.search(index=config.OPENSEARCH_INDEX, body=body, scroll="5m")
    sid = r.get("_scroll_id")
    total = r["hits"]["total"]["value"]
    rows = []
    count = 0

    for hit in r["hits"]["hits"]:
        rows.append(hit["_source"])
        count += 1
        if count >= max_rows:
            break

    while sid and count < max_rows:
        r = client.scroll(scroll_id=sid, scroll="5m")
        if not r["hits"]["hits"]:
            break
        for hit in r["hits"]["hits"]:
            rows.append(hit["_source"])
            count += 1
            if count >= max_rows:
                break
        if progress_cb and total > 0:
            pct = min(90, int(count / total * 90))
            progress_cb(pct, f"{progress_label}: {count:,} / {total:,} rows")

    try:
        client.clear_scroll(scroll_id=sid)
    except:
        pass

    return rows, total


def _get_nested(obj, path, default=""):
    """Get nested dict value by dot path"""
    parts = path.split(".")
    for p in parts:
        if isinstance(obj, dict):
            obj = obj.get(p, default)
        else:
            return default
    return obj if obj is not None else default


def _agg_query(query, aggs):
    """Run aggregation on alerts index"""
    client = opensearch_client.get_client()
    body = {"size": 0, "query": query, "aggs": aggs, "track_total_hits": True}
    return client.search(index=config.OPENSEARCH_INDEX, body=body)


# ============================================================
# EXPORT TYPES
# ============================================================

def export_security_events(period="24h", agent=None, progress_cb=None):
    """Full security events export — matches PDF report structure exactly"""
    # Use the same query as pdf_generator
    import pdf_generator
    query_body = pdf_generator.build_query(period)
    if agent:
        query_body["query"]["bool"] = query_body.get("query", {}).get("bool", {"must": []})
        if "must" not in query_body["query"]["bool"]:
            query_body["query"]["bool"]["must"] = []
        query_body["query"]["bool"]["must"].append({"term": {"agent.name": agent}})

    client = opensearch_client.get_client()
    raw = client.search(index=config.OPENSEARCH_INDEX, body=query_body)
    data = pdf_generator.process_data(raw, period=period)

    time_from = f"now-{period}"
    must = [{"range": {"timestamp": {"gte": time_from}}}]
    if agent:
        must.append({"term": {"agent.name": agent}})
    query = {"bool": {"must": must}}

    wb = Workbook()
    sc = data["severity_counts"]
    ps = data["prev_severity"]

    # Sheet 1: Executive Summary
    if progress_cb: progress_cb(5, "Building executive summary...")
    ws = wb.active
    ws.title = "Executive Summary"
    ws.cell(1, 1, "Security Threat Analysis Report").font = TITLE_FONT
    pl = {"24h": "Daily (Last 24 Hours)", "7d": "Weekly (Last 7 Days)", "30d": "Monthly (Last 30 Days)"}.get(period, period)
    ws.cell(2, 1, f"Period: {pl} | Agent: {agent or 'All'} | Generated: {datetime.now(IST).strftime('%d/%m/%Y %H:%M IST')}").font = SUBTITLE_FONT

    row = 4
    _add_header(ws, row, ["Metric", "Current Period", "Previous Period", "Change"])
    row += 1
    for label, curr, prev in [
        ("Total Alerts", data["total_alerts"], data["prev_total"]),
        ("Critical (15+)", sc["Critical"], ps["Critical"]),
        ("High (12-14)", sc["High"], ps["High"]),
        ("Medium (7-11)", sc["Medium"], ps["Medium"]),
        ("Low (0-6)", sc["Low"], ps["Low"]),
    ]:
        diff = curr - prev
        pct = f"{(diff/prev*100):.1f}%" if prev > 0 else "N/A"
        change = f"+{pct}" if diff > 0 else pct
        _add_row(ws, row, [label, curr, prev, change], row % 2 == 0)
        # Color severity cells
        if label != "Total Alerts":
            sev_key = label.split(" ")[0]
            if sev_key in SEV_COLORS:
                ws.cell(row, 1).fill = SEV_COLORS[sev_key]
                ws.cell(row, 1).font = Font(name='Segoe UI', bold=True, size=9, color='FFFFFF')
        row += 1

    # Get actual agent count (not limited by top_agents size)
    try:
        agent_count_r = _agg_query(query, {"unique_agents": {"cardinality": {"field": "agent.name"}}})
        actual_agent_count = agent_count_r["aggregations"]["unique_agents"]["value"]
    except:
        actual_agent_count = len(data["top_agents"])

    row += 1
    for label, val in [
        ("Auth Failures", data["auth_fail_count"]),
        ("Auth Successes", data["auth_success_count"]),
        ("Level 12+ Events", data["level12_count"]),
        ("Vulnerability Events (alerts)", data["vuln_total"]),
        ("FIM Events", data["fim_total"]),
        ("MITRE Techniques", len(data["mitre_techniques"])),
        ("MITRE Tactics", len(data["mitre_tactics"])),
        ("Active Agents", actual_agent_count),
    ]:
        _add_row(ws, row, [label, val], row % 2 == 0)
        row += 1
    _auto_width(ws)

    # Sheet 2: Top Threat Alerts (Level 12+) — sorted by level desc
    if progress_cb: progress_cb(10, "Building threat alerts...")
    ws2 = wb.create_sheet("Top Threat Alerts")
    _add_header(ws2, 1, ["#", "Level", "Severity", "Rule ID", "Description", "Count", "Affected Agents", "Last Seen"])
    for i, inc in enumerate(data["incidents"]):
        sev_name = "Critical" if inc["level"] >= 15 else "High"
        _add_row(ws2, i+2, [i+1, inc["level"], sev_name, inc["rule_id"], inc["description"],
                             inc["count"], inc["agents"], inc["last_seen"]], i % 2 == 1)
        # Color the severity cell
        fill = SEV_COLORS.get(sev_name)
        if fill:
            ws2.cell(i+2, 3).fill = fill
            ws2.cell(i+2, 3).font = Font(name='Segoe UI', bold=True, size=9, color='FFFFFF')
    _auto_width(ws2)

    # Sheet 3: Agents & Risk Assessment — with severity breakdown
    if progress_cb: progress_cb(15, "Building agents risk...")
    ws3 = wb.create_sheet("Agents & Risk")
    _add_header(ws3, 1, ["#", "Agent", "Total Alerts", "Critical", "High", "Medium", "Low", "Risk Score", "Risk Level", "Action"])
    for i, a in enumerate(data["agent_risk"]):
        _add_row(ws3, i+2, [i+1, a["name"], a["count"], a["critical"], a["high"], a["medium"], a["low"],
                             a["score"], a["risk"], a["note"]], i % 2 == 1)
        # Color risk level
        risk_fill = SEV_COLORS.get(a["risk"])
        if risk_fill:
            ws3.cell(i+2, 9).fill = risk_fill
            ws3.cell(i+2, 9).font = Font(name='Segoe UI', bold=True, size=9, color='FFFFFF')
    _auto_width(ws3)

    # Sheet 4: Authentication — separated sections with headers
    if progress_cb: progress_cb(20, "Building authentication...")
    ws4 = wb.create_sheet("Authentication")
    FAIL_FILL = PatternFill(start_color='FDF2F2', end_color='FDF2F2', fill_type='solid')
    SUCCESS_FILL = PatternFill(start_color='F0FDF4', end_color='F0FDF4', fill_type='solid')
    SECTION_FONT = Font(name='Segoe UI', bold=True, size=12, color='BD271E')
    SECTION_FONT_G = Font(name='Segoe UI', bold=True, size=12, color='2ECC71')

    row = 1
    # --- FAILURES ---
    ws4.cell(row, 1, f"AUTHENTICATION FAILURES ({data['auth_fail_count']:,})").font = SECTION_FONT
    ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    ws4.cell(row, 1, "Failed Login Users").font = Font(name='Segoe UI', bold=True, size=10, color='BD271E')
    row += 1
    _add_header(ws4, row, ["User", "Failure Count"])
    row += 1
    for u in data["auth_fail_users"]:
        _add_row(ws4, row, [u["user"], u["count"]], row % 2 == 0)
        ws4.cell(row, 1).fill = FAIL_FILL
        row += 1

    row += 1
    ws4.cell(row, 1, "Failed Source IPs").font = Font(name='Segoe UI', bold=True, size=10, color='BD271E')
    row += 1
    _add_header(ws4, row, ["IP Address", "Attempts"])
    row += 1
    for ip in data["auth_fail_ips"]:
        _add_row(ws4, row, [ip["ip"], ip["count"]], row % 2 == 0)
        ws4.cell(row, 1).fill = FAIL_FILL
        row += 1

    row += 1
    ws4.cell(row, 1, "Failed Login Events").font = Font(name='Segoe UI', bold=True, size=10, color='BD271E')
    row += 1
    _add_header(ws4, row, ["Rule Description", "Count"])
    row += 1
    for e in data["auth_fail_events"]:
        _add_row(ws4, row, [e["desc"], e["count"]], row % 2 == 0)
        ws4.cell(row, 1).fill = FAIL_FILL
        row += 1

    # --- SUCCESSES ---
    row += 2
    ws4.cell(row, 1, f"AUTHENTICATION SUCCESSES ({data['auth_success_count']:,})").font = SECTION_FONT_G
    ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    ws4.cell(row, 1, "Successful Login Users").font = Font(name='Segoe UI', bold=True, size=10, color='2ECC71')
    row += 1
    _add_header(ws4, row, ["User", "Logon Count"])
    row += 1
    for u in data["auth_success_users"]:
        _add_row(ws4, row, [u["user"], u["count"]], row % 2 == 0)
        ws4.cell(row, 1).fill = SUCCESS_FILL
        row += 1

    row += 1
    ws4.cell(row, 1, "Successful Login Events").font = Font(name='Segoe UI', bold=True, size=10, color='2ECC71')
    row += 1
    _add_header(ws4, row, ["Rule Description", "Count"])
    row += 1
    for e in data["auth_success_events"]:
        _add_row(ws4, row, [e["desc"], e["count"]], row % 2 == 0)
        ws4.cell(row, 1).fill = SUCCESS_FILL
        row += 1
    _auto_width(ws4)

    # Sheet 5: Top Source IPs — separated by Internal/External
    ws5 = wb.create_sheet("Source IPs")
    INT_FILL = PatternFill(start_color='E8F4FD', end_color='E8F4FD', fill_type='solid')
    EXT_FILL = PatternFill(start_color='FDF2F2', end_color='FDF2F2', fill_type='solid')

    # Split IPs into internal and external
    internal_ips = []
    external_ips = []
    for ip in data["top_srcips"]:
        is_priv = ip["ip"].startswith("192.168.") or ip["ip"].startswith("10.") or ip["ip"].startswith("172.") or ip["ip"].startswith("fe80")
        if is_priv:
            internal_ips.append(ip)
        else:
            external_ips.append(ip)

    row = 1
    ws5.cell(row, 1, f"INTERNAL NETWORK ({len(internal_ips)} IPs)").font = Font(name='Segoe UI', bold=True, size=12, color='3498DB')
    row += 2
    _add_header(ws5, row, ["#", "IP Address", "Type", "Events"])
    row += 1
    for i, ip in enumerate(internal_ips):
        _add_row(ws5, row, [i+1, ip["ip"], "Internal", ip["count"]], False)
        ws5.cell(row, 3).fill = INT_FILL
        ws5.cell(row, 3).font = Font(name='Segoe UI', size=9, color='3498DB', bold=True)
        row += 1

    row += 1
    ws5.cell(row, 1, f"EXTERNAL / PUBLIC ({len(external_ips)} IPs)").font = Font(name='Segoe UI', bold=True, size=12, color='E74C3C')
    row += 2
    _add_header(ws5, row, ["#", "IP Address", "Type", "Events"])
    row += 1
    for i, ip in enumerate(external_ips):
        _add_row(ws5, row, [i+1, ip["ip"], "External", ip["count"]], False)
        ws5.cell(row, 3).fill = EXT_FILL
        ws5.cell(row, 3).font = Font(name='Segoe UI', size=9, color='E74C3C', bold=True)
        row += 1
    _auto_width(ws5)

    # Sheet 6: Vulnerability
    if progress_cb: progress_cb(25, "Building vulnerability...")
    ws6 = wb.create_sheet("Vulnerability")
    ws6.cell(1, 1, f"Total Vulnerability Events: {data['vuln_total']:,}").font = TITLE_FONT
    row = 3
    if data["vuln_by_sev"]:
        _add_header(ws6, row, ["Severity", "Count"])
        row += 1
        for v in data["vuln_by_sev"]:
            _add_row(ws6, row, [v["severity"], v["count"]], row % 2 == 0)
            fill = SEV_COLORS.get(v["severity"])
            if fill:
                ws6.cell(row, 1).fill = fill
                ws6.cell(row, 1).font = Font(name='Segoe UI', bold=True, size=9, color='FFFFFF')
            row += 1
    row += 1
    if data["vuln_top_cve"]:
        ws6.cell(row, 1, "Top CVEs").font = TITLE_FONT
        row += 1
        _add_header(ws6, row, ["CVE ID", "Count"])
        row += 1
        for c in data["vuln_top_cve"]:
            _add_row(ws6, row, [c["cve"], c["count"]], row % 2 == 0)
            row += 1
    row += 1
    if data["vuln_top_agent"]:
        ws6.cell(row, 1, "Affected Agents").font = TITLE_FONT
        row += 1
        _add_header(ws6, row, ["Agent", "Vulns"])
        row += 1
        for a in data["vuln_top_agent"]:
            _add_row(ws6, row, [a["agent"], a["count"]], row % 2 == 0)
            row += 1
    _auto_width(ws6)

    # Sheet 7: File Integrity Monitoring
    ws7 = wb.create_sheet("FIM")
    ws7.cell(1, 1, f"Total FIM Events: {data['fim_total']:,}").font = TITLE_FONT
    row = 3
    if data["fim_events"]:
        _add_header(ws7, row, ["Event Type", "Count"])
        row += 1
        for e in data["fim_events"]:
            _add_row(ws7, row, [e["event"], e["count"]], row % 2 == 0)
            row += 1
    row += 1
    if data["fim_agents"]:
        ws7.cell(row, 1, "FIM by Agent").font = TITLE_FONT
        row += 1
        _add_header(ws7, row, ["Agent", "Changes"])
        row += 1
        for a in data["fim_agents"]:
            _add_row(ws7, row, [a["agent"], a["count"]], row % 2 == 0)
            row += 1
    row += 1
    if data["fim_paths"]:
        ws7.cell(row, 1, "Top Modified Paths").font = TITLE_FONT
        row += 1
        _add_header(ws7, row, ["Path", "Count"])
        row += 1
        for p in data["fim_paths"]:
            _add_row(ws7, row, [p["path"], p["count"]], row % 2 == 0)
            row += 1
    _auto_width(ws7)

    # Sheet 8: MITRE ATT&CK
    ws8 = wb.create_sheet("MITRE ATT&CK")
    _add_header(ws8, 1, ["#", "Technique", "Count"])
    for i, t in enumerate(data["mitre_techniques"]):
        _add_row(ws8, i+2, [i+1, t["technique"], t["count"]], i % 2 == 1)
    row = len(data["mitre_techniques"]) + 3
    ws8.cell(row, 1, "Tactics").font = TITLE_FONT
    row += 1
    _add_header(ws8, row, ["#", "Tactic", "Count"])
    row += 1
    for i, t in enumerate(data["mitre_tactics"]):
        _add_row(ws8, row, [i+1, t["tactic"], t["count"]], i % 2 == 1)
        row += 1
    _auto_width(ws8)

    # Sheet 9: Compliance
    if progress_cb: progress_cb(30, "Building compliance...")
    ws9 = wb.create_sheet("Compliance")
    comp = data.get("compliance", {})
    row = 1
    for fw_name, fw_key in [("PCI-DSS", "pci"), ("HIPAA", "hipaa"), ("GDPR", "gdpr"), ("NIST 800-53", "nist"), ("TSC", "tsc")]:
        fw = comp.get(fw_key, {})
        ws9.cell(row, 1, fw_name).font = TITLE_FONT
        ws9.cell(row, 2, f"Total: {fw.get('total', 0):,}").font = SUBTITLE_FONT
        row += 1
        _add_header(ws9, row, ["Control", "Alerts"])
        row += 1
        for c in fw.get("controls", []):
            _add_row(ws9, row, [c["control"], c["count"]], row % 2 == 0)
            row += 1
        row += 1
    _auto_width(ws9)

    # Sheet 10: All Rules (sorted by level desc, then count desc) — with severity colors
    # Re-query to get ALL rules (not just top 50 by count) — sorted by level desc
    if progress_cb: progress_cb(35, "Building security events summary...")
    ws10 = wb.create_sheet("Security Events")
    try:
        all_rules_r = _agg_query(query, {
            "all_rules": {"terms": {"field": "rule.id", "size": 500, "order": {"max_level": "desc"}},
                "aggs": {"desc": {"terms": {"field": "rule.description", "size": 1}},
                         "max_level": {"max": {"field": "rule.level"}}}}
        })
        sorted_rules = []
        for b in all_rules_r["aggregations"]["all_rules"]["buckets"]:
            desc = b["desc"]["buckets"][0]["key"] if b["desc"]["buckets"] else "N/A"
            lvl = int(b["max_level"]["value"]) if b["max_level"]["value"] else 0
            sorted_rules.append({"rule_id": b["key"], "description": desc, "level": lvl, "count": b["doc_count"]})
        sorted_rules.sort(key=lambda x: (-x["level"], -x["count"]))
    except:
        sorted_rules = sorted(data["top_rules"], key=lambda x: (-x["level"], -x["count"]))
    _add_header(ws10, 1, ["#", "Rule ID", "Description", "Level", "Severity", "Count"])
    for i, r in enumerate(sorted_rules):
        sev = _severity(r["level"]) if isinstance(r["level"], int) else "Low"
        _add_row(ws10, i+2, [i+1, r["rule_id"], r["description"], r["level"], sev, r["count"]], i % 2 == 1)
        fill = SEV_COLORS.get(sev)
        if fill:
            ws10.cell(i+2, 5).fill = fill
            ws10.cell(i+2, 5).font = Font(name='Segoe UI', bold=True, size=9, color='FFFFFF')
    _auto_width(ws10)

    # Sheet 11: Recommendations
    ws11 = wb.create_sheet("Recommendations")
    _add_header(ws11, 1, ["#", "Priority", "Recommendation"])
    for i, r in enumerate(data["recommendations"]):
        _add_row(ws11, i+2, [i+1, r["icon"], r["text"]], i % 2 == 1)
    _auto_width(ws11)

    # Save
    if progress_cb: progress_cb(95, "Saving Excel file...")
    period_label = {"24h": "Daily", "7d": "Weekly", "30d": "Monthly", "90d": "Quarterly"}.get(period, period)
    agent_label = agent.replace(" ", "_") if agent else "All_Agents"
    fname = f"security_events_{agent_label}_{period_label}_{datetime.now(IST).strftime('%Y%m%d_%H%M')}.xlsx"
    fpath = os.path.join(config.REPORTS_DIR, fname)
    os.makedirs(config.REPORTS_DIR, exist_ok=True)
    wb.save(fpath)

    size = os.path.getsize(fpath)
    return {"filename": fname, "filepath": fpath, "size": size, "total_events": data["total_alerts"],
            "sheets": ["Executive Summary", "Top Threat Alerts", "Agents & Risk", "Authentication",
                       "Source IPs", "Vulnerability", "FIM", "MITRE ATT&CK", "Compliance", "Security Events", "Recommendations"]}


def export_auth_events(period="24h", agent=None, progress_cb=None):
    """Authentication events Excel — failed/success logins with details"""
    time_from = f"now-{period}"
    must = [{"range": {"timestamp": {"gte": time_from}}},
            {"bool": {"should": [
                {"term": {"rule.groups": "win_authentication_failed"}},
                {"term": {"rule.groups": "authentication_failed"}},
                {"term": {"rule.groups": "authentication_failures"}},
                {"term": {"rule.groups": "authentication_success"}}
            ], "minimum_should_match": 1}}]
    if agent:
        must.append({"term": {"agent.name": agent}})
    query = {"bool": {"must": must}}

    wb = Workbook()
    ws = wb.active
    ws.title = "Auth Events"

    fields = ["timestamp", "agent.name", "rule.id", "rule.description", "rule.level", "rule.groups",
              "data.win.eventdata.targetUserName", "data.win.eventdata.ipAddress",
              "data.dstuser", "data.srcuser", "data.srcip"]
    headers = ["Timestamp", "Agent", "Rule ID", "Description", "Level", "Groups",
               "Target User (Win)", "Source IP (Win)", "Dest User", "Source User", "Source IP"]

    if progress_cb: progress_cb(10, "Pulling auth events...")
    _add_header(ws, 1, headers)
    rows_data, total = _scroll_alerts(query, fields, MAX_ROWS, progress_cb, "Auth events")

    for i, src in enumerate(rows_data):
        values = [
            _get_nested(src, "timestamp"), _get_nested(src, "agent.name"),
            _get_nested(src, "rule.id"), _get_nested(src, "rule.description"),
            _get_nested(src, "rule.level", 0),
            ", ".join(_get_nested(src, "rule.groups", [])) if isinstance(_get_nested(src, "rule.groups", []), list) else "",
            _get_nested(src, "data.win.eventdata.targetUserName"),
            _get_nested(src, "data.win.eventdata.ipAddress"),
            _get_nested(src, "data.dstuser"), _get_nested(src, "data.srcuser"),
            _get_nested(src, "data.srcip")
        ]
        _add_row(ws, i+2, values, i % 2 == 1)

    _auto_width(ws)

    if progress_cb: progress_cb(95, "Saving...")
    agent_label = agent.replace(" ", "_") if agent else "All_Agents"
    fname = f"auth_events_{agent_label}_{datetime.now(IST).strftime('%Y%m%d_%H%M')}.xlsx"
    fpath = os.path.join(config.REPORTS_DIR, fname)
    wb.save(fpath)
    return {"filename": fname, "filepath": fpath, "size": os.path.getsize(fpath), "total_events": total, "sheets": ["Auth Events"]}


def export_vulnerability(agent=None, progress_cb=None):
    """Vulnerability data Excel — from wazuh-states-vulnerabilities-*"""
    client = opensearch_client.get_client()
    idx = config.OPENSEARCH_INDEX.split('-')[0] + "-states-vulnerabilities-*"

    wb = Workbook()
    ws = wb.active
    ws.title = "Vulnerabilities"

    headers = ["Agent", "CVE ID", "Severity", "CVSS Score", "Description", "Package", "Package Version",
               "OS", "Category", "Detected At", "Published At", "Reference"]
    _add_header(ws, 1, headers)

    query = {"match_all": {}}
    if agent:
        query = {"term": {"agent.name": agent}}

    body = {"size": 5000, "query": query, "sort": [{"vulnerability.score.base": "desc"}]}
    r = client.search(index=idx, body=body, scroll="5m")
    sid = r.get("_scroll_id")
    total = r["hits"]["total"]["value"]
    row = 2
    count = 0

    def write_hits(hits):
        nonlocal row, count
        for hit in hits:
            s = hit["_source"]
            values = [
                _get_nested(s, "agent.name"), _get_nested(s, "vulnerability.id"),
                _get_nested(s, "vulnerability.severity"), _get_nested(s, "vulnerability.score.base", 0),
                _get_nested(s, "vulnerability.description"),
                _get_nested(s, "package.name"), _get_nested(s, "package.version"),
                _get_nested(s, "host.os.name"), _get_nested(s, "vulnerability.category"),
                _get_nested(s, "vulnerability.detected_at"), _get_nested(s, "vulnerability.published_at"),
                _get_nested(s, "vulnerability.reference")
            ]
            _add_row(ws, row, values, row % 2 == 0)
            row += 1
            count += 1
            if count >= MAX_ROWS:
                return

    write_hits(r["hits"]["hits"])
    while sid and count < MAX_ROWS:
        r = client.scroll(scroll_id=sid, scroll="5m")
        if not r["hits"]["hits"]:
            break
        write_hits(r["hits"]["hits"])
        if progress_cb:
            pct = min(90, int(count / max(total, 1) * 90))
            progress_cb(pct, f"Vulnerabilities: {count:,} / {total:,}")

    try:
        client.clear_scroll(scroll_id=sid)
    except:
        pass

    _auto_width(ws)

    if progress_cb: progress_cb(95, "Saving...")
    agent_label = agent.replace(" ", "_") if agent else "All_Agents"
    fname = f"vulnerabilities_{agent_label}_{datetime.now(IST).strftime('%Y%m%d_%H%M')}.xlsx"
    fpath = os.path.join(config.REPORTS_DIR, fname)
    wb.save(fpath)
    return {"filename": fname, "filepath": fpath, "size": os.path.getsize(fpath), "total_events": total, "sheets": ["Vulnerabilities"]}
